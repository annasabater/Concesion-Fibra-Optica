"""Carga y normalización de los datos de partida del proyecto ABAST.

Este módulo es la **única** capa que toca disco. El resto del proyecto
trabaja con DataFrames y dicts en memoria devueltos por estas funciones.

Funciones públicas:
    load_topology(path) -> pd.DataFrame
    load_parameters(path) -> dict
    load_decisiones(path, escenario="base") -> dict
"""

from __future__ import annotations

import difflib
import logging
import re
import unicodedata
from copy import deepcopy
from pathlib import Path

import openpyxl
import pandas as pd
import yaml

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constantes y helpers de normalización
# ---------------------------------------------------------------------------

_TOPOLOGY_SHEET = "Municipios y topologia de red"
_PARAMETERS_SHEET = "Parámetros"

_RING_AGG_PATTERN = re.compile(r"^A([1-9]|1[01])$")
_RING_TRO_PATTERN = re.compile(r"^T[1-3]$")

# Convención del Excel: 1–799 acceso, 800–890 agregación, 891–900 troncal.
_TIER_CODE_RANGES: dict[str, tuple[int, int]] = {
    "acceso":     (1, 799),
    "agregacion": (800, 890),
    "troncal":    (891, 900),
}

# Cabeceras del Excel (ya normalizadas) → nombres canónicos en el DataFrame.
# Incluye la variante con la errata 'agrefacion' que aparece en la hoja real.
_TOPOLOGY_HEADER_MAP = {
    "codigo": "codigo",
    "hab": "hab",
    "sedes_abast": "sedes_abast",
    "num_operadores": "num_operadores",
    "municipio_origen_servicio": "municipio",
    "municipio_destino_acceso": "destino_acceso",
    "km": "km",
    "municipio_destino_agregacion": "destino_agregacion",
    "municipio_destino_agrefacion": "destino_agregacion",
    "municipio_destino_troncal": "destino_troncal",
    "anillo_agregacion": "anillo_agregacion",
    "anillo_troncal": "anillo_troncal",
}

_OUTPUT_COLUMNS = [
    "codigo", "hab", "sedes_abast", "num_operadores", "municipio",
    "destino_acceso", "km", "anillo_agregacion", "anillo_troncal", "tier",
]

# Claves obligatorias que se validan tras el deep_merge en load_decisiones.
# Cada entrada es una ruta dot-separated dentro del dict del escenario.
_REQUIRED_DECISIONES_KEYS: tuple[str, ...] = (
    "mayorista.cuota_mercado.tipo",
    "mayorista.bw_por_hogar_mbps",
    "mayorista.habitantes_por_hogar",
    "mayorista.overbooking",
    "abast.bw_por_sede_mbps",
    "equipos.umbral_10p_a_20p",
    "despliegue.estrategia",
    "descuento.tasa",
    "descuento.anios",
)


def _normalize(value: object) -> str | None:
    """Quita tildes, baja a minúsculas y compacta espacios → snake_case."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    nfkd = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    s = s.lower().replace("-", " ").replace("/", " ")
    s = re.sub(r"\s+", "_", s.strip())
    return s


def _is_na(value: object) -> bool:
    """True para None, NaN, cadena vacía o el literal 'NA'/'N/A'."""
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip().upper() in {"NA", "N/A", ""}:
        return True
    return False


def _normalize_ring(value: object, pattern: re.Pattern) -> str | None:
    """Limpia 'T1 ', 'T1, T2,T3' o 'A1, A2' → cadena ordenada y validada.

    Devuelve None si la celda es NA. Si hay varios anillos, se ordenan por
    índice numérico y se serializan separados por coma sin espacios
    (ej. 'T1, T2,T3' → 'T1,T2,T3').
    """
    if _is_na(value):
        return None
    s = str(value).strip().upper()
    parts = [p.strip() for p in re.split(r"[,\s]+", s) if p.strip()]
    if not parts:
        return None
    for p in parts:
        if not pattern.match(p):
            raise ValueError(f"Anillo no reconocido: {value!r}")
    parts_sorted = sorted(parts, key=lambda x: int(x[1:]))
    return ",".join(parts_sorted)


def _detect_header_row(ws) -> int:
    """Localiza la fila de cabecera buscando 'Código' (con o sin tilde)."""
    limit = min(ws.max_row, 20)
    for r in range(1, limit + 1):
        for c in range(1, ws.max_column + 1):
            if _normalize(ws.cell(row=r, column=c).value) == "codigo":
                return r
    raise ValueError("No se encontró la fila de cabecera 'Código' en la hoja.")


def _resolve_sheet(wb, target: str):
    """Devuelve la hoja con el nombre dado, tolerando variaciones de tildes."""
    if target in wb.sheetnames:
        return wb[target]
    norm_target = _normalize(target)
    for sn in wb.sheetnames:
        if _normalize(sn) == norm_target:
            return wb[sn]
    raise ValueError(
        f"Hoja '{target}' no encontrada. Hojas disponibles: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# load_topology
# ---------------------------------------------------------------------------

def load_topology(path: Path) -> pd.DataFrame:
    """Carga la hoja 'Municipios y topologia de red' del Excel.

    Args:
        path: ruta al archivo Excel con la hoja de topología.

    Returns:
        DataFrame con columnas:
            codigo (int), hab (int), sedes_abast (int), num_operadores (int),
            municipio (str), destino_acceso (str | None), km (float),
            anillo_agregacion (str | None), anillo_troncal (str | None),
            tier (str: 'acceso' | 'agregacion' | 'troncal').

    El campo `tier` se calcula:
        - si anillo_troncal != NA → 'troncal'
        - elif anillo_agregacion != NA → 'agregacion'
        - else → 'acceso'

    Para los nodos de agregación/troncal (códigos 800–900), el
    `destino_acceso` original puede contener fórmulas autorreferenciales;
    se normaliza a None porque son terminadores del árbol.

    Formato de las columnas de anillo:
        `anillo_agregacion` y `anillo_troncal` son cadenas normalizadas
        sin espacios y ordenadas numéricamente. Casos multi-anillo se
        serializan separados por coma. Ejemplos:
            'T1'        → un único anillo troncal
            'T1,T2,T3'  → A900 (en los 3 anillos troncales)
            'A1,A2'     → A900 (en agregación A1 y A2)

        Para reconstruir un `set` de anillos en `topology.py`:
            anillos = set(s.split(',')) if s else set()

    Raises:
        ValueError: si falta la hoja, faltan columnas, o los datos infringen
            las validaciones (códigos duplicados/incompletos, rangos, anillos
            no reconocidos, o inconsistencia tier↔rango de código).
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = _resolve_sheet(wb, _TOPOLOGY_SHEET)
    header_row = _detect_header_row(ws)

    # Mapear índices de columna (1-based) → nombre canónico.
    columns: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        norm = _normalize(ws.cell(row=header_row, column=c).value)
        if norm in _TOPOLOGY_HEADER_MAP:
            columns[c] = _TOPOLOGY_HEADER_MAP[norm]

    expected = {
        "codigo", "hab", "sedes_abast", "num_operadores", "municipio",
        "destino_acceso", "km", "anillo_agregacion", "anillo_troncal",
    }
    missing = expected - set(columns.values())
    if missing:
        raise ValueError(f"Faltan columnas esperadas en la hoja: {sorted(missing)}")

    # iter_rows(values_only=True) devuelve tuplas alineadas con min_col=1,
    # así que la posición i de la tupla corresponde a la columna i+1.
    position_to_name = {col - 1: name for col, name in columns.items()}
    codigo_pos = next(p for p, name in position_to_name.items() if name == "codigo")

    rows: list[dict] = []
    for row_values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        codigo_raw = row_values[codigo_pos] if codigo_pos < len(row_values) else None
        if codigo_raw is None or (isinstance(codigo_raw, str) and not codigo_raw.strip()):
            continue
        record = {
            name: (row_values[pos] if pos < len(row_values) else None)
            for pos, name in position_to_name.items()
        }
        rows.append(record)

    df = pd.DataFrame(rows)

    # Coerciones de tipo.
    df["codigo"] = pd.to_numeric(df["codigo"], errors="raise").astype(int)
    df["hab"] = pd.to_numeric(df["hab"], errors="raise").astype(int)
    df["sedes_abast"] = pd.to_numeric(df["sedes_abast"], errors="raise").astype(int)
    df["num_operadores"] = pd.to_numeric(df["num_operadores"], errors="raise").astype(int)
    df["km"] = pd.to_numeric(df["km"], errors="raise").astype(float)
    df["municipio"] = df["municipio"].astype(str).str.strip()

    df["destino_acceso"] = df["destino_acceso"].map(
        lambda v: None if _is_na(v) else str(v).strip()
    )
    df["anillo_agregacion"] = df["anillo_agregacion"].map(
        lambda v: _normalize_ring(v, _RING_AGG_PATTERN)
    )
    df["anillo_troncal"] = df["anillo_troncal"].map(
        lambda v: _normalize_ring(v, _RING_TRO_PATTERN)
    )

    # Códigos 800–900: terminadores del árbol — descartar destino_acceso.
    df.loc[df["codigo"] >= 800, "destino_acceso"] = None

    df["tier"] = df.apply(_compute_tier, axis=1).astype(str)

    df = df[_OUTPUT_COLUMNS].sort_values("codigo").reset_index(drop=True)
    _validate_topology(df)

    logger.info(
        "load_topology: %d filas (acceso=%d, agregacion=%d, troncal=%d), "
        "total sedes_abast=%d, total habitantes=%d",
        len(df),
        int((df["tier"] == "acceso").sum()),
        int((df["tier"] == "agregacion").sum()),
        int((df["tier"] == "troncal").sum()),
        int(df["sedes_abast"].sum()),
        int(df["hab"].sum()),
    )
    return df


def _compute_tier(row: pd.Series) -> str:
    # `map` puede devolver None o NaN según el dtype; usar pd.notna por seguridad.
    if pd.notna(row["anillo_troncal"]):
        return "troncal"
    if pd.notna(row["anillo_agregacion"]):
        return "agregacion"
    return "acceso"


def _validate_topology(df: pd.DataFrame) -> None:
    """Aplica las validaciones documentadas en `prompts/01_load.md`."""
    if df["codigo"].duplicated().any():
        dups = df.loc[df["codigo"].duplicated(), "codigo"].tolist()
        raise ValueError(f"`codigo` con duplicados: {dups[:10]}")

    expected_codes = set(range(1, 901))
    actual = set(df["codigo"].tolist())
    if actual != expected_codes:
        faltan = sorted(expected_codes - actual)
        sobran = sorted(actual - expected_codes)
        raise ValueError(
            f"`codigo` debe cubrir 1–900 sin huecos. "
            f"Faltan={faltan[:10]}, sobran={sobran[:10]}"
        )

    if (df["hab"] < 0).any():
        raise ValueError("`hab` no puede ser negativo.")
    if (df["sedes_abast"] < 0).any():
        raise ValueError("`sedes_abast` no puede ser negativo.")
    if not df["num_operadores"].between(1, 5).all():
        bad = df.loc[~df["num_operadores"].between(1, 5), "codigo"].tolist()
        raise ValueError(f"`num_operadores` fuera de [1,5] en códigos: {bad[:10]}")
    if (df["km"] < 0).any():
        raise ValueError("`km` no puede ser negativo.")

    huecos = df[(df["tier"] == "acceso") & df["destino_acceso"].isna()]
    if not huecos.empty:
        logger.warning(
            "Hay %d municipios de acceso sin destino_acceso (posible agujero "
            "en el árbol). Primeros códigos: %s",
            len(huecos), huecos["codigo"].head(10).tolist(),
        )

    _validate_tier_consistency(df)


def _validate_tier_consistency(df: pd.DataFrame) -> None:
    """Confirma que código 1–799=acceso, 800–890=agregación, 891–900=troncal."""
    for tier, (lo, hi) in _TIER_CODE_RANGES.items():
        sub = df[df["codigo"].between(lo, hi)]
        bad = sub[sub["tier"] != tier]
        if not bad.empty:
            raise ValueError(
                f"Inconsistencia tier↔código: códigos {lo}–{hi} deberían "
                f"tener tier='{tier}'. Filas violatorias: "
                f"{bad[['codigo', 'tier']].head(10).to_dict(orient='records')}"
            )


# ---------------------------------------------------------------------------
# load_parameters
# ---------------------------------------------------------------------------

# Etiquetas normalizadas de la hoja 'Parámetros' → (categoría, clave de salida).
_PARAM_LABEL_MAP: dict[str, tuple[str, str]] = {
    # Ingresos
    "alta_por_sede_de_abast": ("ingresos", "alta_abast"),
    "recurrente_mensual_por_sede_abast": ("ingresos", "recurrente_abast"),
    "alta_por_sede_mayorista": ("ingresos", "alta_mayorista"),
    "recurrente_mensual_por_sede_mayorista": ("ingresos", "recurrente_mayorista"),
    # OPEX
    "mantenimiento_de_fibra_por_km_y_ano": ("opex", "mant_fibra_eur_km_año"),
    "mantenimiento_de_los_equipos_activos_%_del_capex_(inversiones)": (
        "opex", "mant_equipos_pct_capex"
    ),
    "mantenimiento_de_los_sistema_de_informacion_%_del_capex": (
        "opex", "mant_si_pct_capex"
    ),
    "derechos_de_paso_y_impuestos_en_%_de_ingresos": (
        "opex", "derechos_paso_pct_ingresos"
    ),
    "costes_de_ventas_mayoristas_en_%_sobre_ingresos": (
        "opex", "costes_ventas_mayorista_pct"
    ),
    "gastos_generales_por_ano": ("opex", "gastos_generales_anuales"),
    # CAPEX
    "inversion_por_nueva_sede_abast": ("capex", "alta_sede_abast"),
    "inversion_por_nueva_sede_mayorista": ("capex", "alta_sede_mayorista"),
    "infraestructura_obra_civil_+_cable_zona_urbana": (
        "capex", "obra_civil_urbana_eur_m"
    ),
    "infraestructura_obra_civil_+_cable_zona_acceso": (
        "capex", "obra_civil_acceso_eur_m"
    ),
    "nodo_agregacion_troncal_sin_equipos_de_red": ("capex", "nodo_chasis"),
    "equipo_cliente": ("capex", "equipo_cliente"),
    "equipo_agregacion_ethernet_10_puertos_10_100_1000": (
        "capex", "equipo_agreg_10p"
    ),
    "equipo_agregacion_ethernet_20_puertos_10_100_1000": (
        "capex", "equipo_agreg_20p"
    ),
    "equipo_agregacion_ethernet_40_puertos_10_100_1000": (
        "capex", "equipo_agreg_40p"
    ),
    "equipo_troncal_mpls": ("capex", "equipo_troncal_mpls"),
    "equipo_troncal_optico_40_lambdas": ("capex", "equipo_troncal_optico_40l"),
}


def load_parameters(path: Path) -> dict:
    """Carga la hoja 'Parámetros' a un dict anidado por categoría.

    Args:
        path: ruta al archivo Excel con la hoja 'Parámetros'.

    Returns:
        dict con tres categorías {ingresos, opex, capex}; cada una mapea
        clave canónica → valor numérico (€, €/km·año, fracción, …).

    Si una etiqueta del Excel no hace match exacto con `_PARAM_LABEL_MAP`,
    se intenta un match aproximado con `difflib.get_close_matches`
    (cutoff=0.85) y se loguea un warning indicando la aproximación usada.

    Raises:
        ValueError: si falta la hoja o algún parámetro esperado.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = _resolve_sheet(wb, _PARAMETERS_SHEET)

    result: dict[str, dict] = {"ingresos": {}, "opex": {}, "capex": {}}

    for r in range(1, ws.max_row + 1):
        label_value: str | None = None
        numeric_value: float | int | None = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and label_value is None:
                label_value = v
            elif isinstance(v, (int, float)) and not isinstance(v, bool) and numeric_value is None:
                numeric_value = v
        if label_value is None or numeric_value is None:
            continue
        norm_label = _normalize(label_value)
        mapping = _PARAM_LABEL_MAP.get(norm_label)
        if mapping is None:
            close = difflib.get_close_matches(
                norm_label or "",
                list(_PARAM_LABEL_MAP.keys()),
                n=1,
                cutoff=0.85,
            )
            if close:
                logger.warning(
                    "Parámetro '%s' (norm='%s') no es match exacto; "
                    "usando aproximación '%s' (cutoff=0.85).",
                    label_value, norm_label, close[0],
                )
                mapping = _PARAM_LABEL_MAP[close[0]]
            else:
                continue
        cat, key = mapping
        result[cat][key] = numeric_value

    missing = [
        f"{cat}.{key}"
        for (cat, key) in _PARAM_LABEL_MAP.values()
        if key not in result[cat]
    ]
    if missing:
        raise ValueError(f"Parámetros no encontrados en la hoja: {missing}")
    return result


# ---------------------------------------------------------------------------
# load_decisiones
# ---------------------------------------------------------------------------

def load_decisiones(path: Path, escenario: str = "base") -> dict:
    """Carga decisiones.yaml y devuelve el escenario solicitado.

    Si `escenario != 'base'` se hace deep merge sobre la configuración de
    'base', de forma que los escenarios optimista/pesimista heredan los
    valores que no sobrescriben explícitamente. Tras el merge se valida que
    estén presentes todas las claves obligatorias listadas en
    `_REQUIRED_DECISIONES_KEYS`.

    Args:
        path: ruta a decisiones.yaml.
        escenario: nombre del escenario a devolver. Por defecto 'base'.

    Returns:
        dict con la configuración resultante del escenario.

    Raises:
        ValueError: si el escenario no existe en el YAML, o si tras el merge
            falta alguna clave obligatoria.
    """
    path = Path(path)
    with path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    escenarios = data.get("escenarios", {})
    if escenario not in escenarios:
        raise ValueError(
            f"Escenario '{escenario}' no existe. "
            f"Disponibles: {sorted(escenarios)}"
        )
    base = deepcopy(escenarios.get("base", {}))
    if escenario == "base":
        cfg = base
    else:
        cfg = _deep_merge(base, deepcopy(escenarios[escenario]))

    _validate_decisiones(cfg, escenario)
    return cfg


def _deep_merge(base: dict, override: dict) -> dict:
    """Mezcla recursiva: los valores de `override` ganan; los sub-dicts se fusionan."""
    out = dict(base)
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def _validate_decisiones(cfg: dict, escenario: str) -> None:
    """Verifica que todas las claves obligatorias existen tras el deep_merge."""
    missing: list[str] = []
    for path in _REQUIRED_DECISIONES_KEYS:
        node: object = cfg
        for part in path.split("."):
            if isinstance(node, dict) and part in node:
                node = node[part]
            else:
                missing.append(path)
                break
    if missing:
        raise ValueError(
            f"Faltan claves obligatorias en decisiones (escenario='{escenario}'): "
            f"{missing}"
        )
